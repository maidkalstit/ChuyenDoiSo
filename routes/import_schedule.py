"""
routes/import_schedule.py - API endpoints cho import lịch từ file
"""

from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
from models import get_db
from utils.auth import token_required
from utils.file_parser import (
    parse_excel_schedule,
    parse_csv_schedule,
    validate_schedules,
    allowed_file,
    get_file_extension
)
from utils.conflict_detector import detect_schedule_conflicts
import os
from datetime import datetime

import_bp = Blueprint('import', __name__)

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


@import_bp.route('/upload', methods=['POST'])
@token_required
def upload_file(current_user):
    """
    Upload file Excel/CSV để import lịch
    
    Form Data:
        file: Excel/CSV file
        preview: true/false (nếu true chỉ preview, không import)
    """
    user_id = current_user['user_id']
    
    # Check if file is in request
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    # Validate file extension
    if not allowed_file(file.filename, ALLOWED_EXTENSIONS):
        return jsonify({
            'error': f'Invalid file type. Allowed: {", ".join(ALLOWED_EXTENSIONS)}'
        }), 400
    
    # Check file size
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    
    if file_size > MAX_FILE_SIZE:
        return jsonify({
            'error': f'File too large. Maximum size: {MAX_FILE_SIZE / 1024 / 1024}MB'
        }), 400
    
    # Save file
    filename = secure_filename(file.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_filename = f"{user_id}_{timestamp}_{filename}"
    file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
    
    try:
        file.save(file_path)
    except Exception as e:
        return jsonify({'error': f'Failed to save file: {str(e)}'}), 500
    
    # Parse file
    file_ext = get_file_extension(filename)
    
    if file_ext in ['xlsx', 'xls']:
        result = parse_excel_schedule(file_path)
    elif file_ext == 'csv':
        result = parse_csv_schedule(file_path)
    else:
        os.remove(file_path)
        return jsonify({'error': 'Unsupported file format'}), 400
    
    # Check for parsing errors
    if 'error' in result:
        os.remove(file_path)
        return jsonify({'error': result['error']}), 400
    
    schedules = result['schedules']
    parsing_errors = result.get('errors', [])
    
    # Validate schedules
    valid_schedules, invalid_schedules, warnings = validate_schedules(schedules)
    
    # Check for conflicts
    conflicts_summary = []
    for schedule in valid_schedules:
        conflicts = detect_schedule_conflicts(
            user_id,
            schedule['start_time'],
            schedule['end_time']
        )
        if conflicts:
            conflicts_summary.append({
                'schedule': schedule['subject'],
                'time': schedule['start_time'],
                'conflicts_with': [c['subject'] for c in conflicts]
            })
    
    # Preview mode
    is_preview = request.form.get('preview', 'false').lower() == 'true'
    
    if is_preview:
        os.remove(file_path)  # Clean up
        
        return jsonify({
            'message': 'File parsed successfully (preview mode)',
            'total_rows': len(schedules),
            'valid_count': len(valid_schedules),
            'invalid_count': len(invalid_schedules),
            'conflicts_count': len(conflicts_summary),
            'preview': valid_schedules[:10],  # Show first 10
            'invalid': invalid_schedules,
            'conflicts': conflicts_summary,
            'warnings': warnings,
            'parsing_errors': parsing_errors
        }), 200
    
    # Import mode - save to database
    imported_count = 0
    failed_imports = []
    
    with get_db() as conn:
        cursor = conn.cursor()
        
        for schedule in valid_schedules:
            try:
                cursor.execute('''
                    INSERT INTO schedule 
                    (user_id, subject, description, start_time, end_time, 
                     location, type, color, reminder_time)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    user_id,
                    schedule['subject'],
                    schedule.get('description', ''),
                    schedule['start_time'],
                    schedule['end_time'],
                    schedule.get('location', ''),
                    schedule.get('type', 'class'),
                    schedule.get('color', '#3788d8'),
                    schedule.get('reminder_time', 30)
                ))
                imported_count += 1
            except Exception as e:
                failed_imports.append({
                    'schedule': schedule['subject'],
                    'error': str(e)
                })
    
    # Clean up file
    os.remove(file_path)
    
    return jsonify({
        'message': 'Import completed',
        'imported_count': imported_count,
        'failed_count': len(failed_imports),
        'invalid_count': len(invalid_schedules),
        'conflicts_count': len(conflicts_summary),
        'failed_imports': failed_imports,
        'invalid_schedules': invalid_schedules,
        'conflicts': conflicts_summary,
        'warnings': warnings
    }), 200


@import_bp.route('/template/excel', methods=['GET'])
def download_excel_template():
    """
    Download template Excel để người dùng điền
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule Template"
        
        # Headers
        headers = ['Môn học', 'Ngày', 'Giờ bắt đầu', 'Giờ kết thúc', 'Phòng', 'Loại', 'Mô tả']
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Add sample data
        sample_data = [
            ['Toán Cao Cấp', '2025-10-20', '08:00', '10:00', 'A101', 'class', 'Chương 3: Tích phân'],
            ['Lập Trình Python', '2025-10-20', '14:00', '16:00', 'Lab 205', 'class', 'Thực hành Flask'],
            ['Kiểm tra giữa kỳ', '2025-10-25', '09:00', '11:00', 'B302', 'exam', 'Thi trắc nghiệm']
        ]
        
        for row_idx, data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='schedule_template.xlsx'
        )
    
    except ImportError:
        return jsonify({
            'error': 'openpyxl not installed. Install with: pip install openpyxl'
        }), 500
    except Exception as e:
        return jsonify({'error': f'Failed to generate template: {str(e)}'}), 500


@import_bp.route('/template/csv', methods=['GET'])
def download_csv_template():
    """Download template CSV"""
    import csv
    from io import StringIO
    from flask import Response
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow(['subject', 'date', 'start_time', 'end_time', 'location', 'type', 'description'])
    
    # Sample data
    writer.writerow(['Toán Cao Cấp', '2025-10-20', '08:00', '10:00', 'A101', 'class', 'Chương 3'])
    writer.writerow(['Lập Trình Python', '2025-10-20', '14:00', '16:00', 'Lab 205', 'class', 'Thực hành'])
    
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment;filename=schedule_template.csv'}
    )


@import_bp.route('/history', methods=['GET'])
@token_required
def get_import_history(current_user):
    """
    Lấy lịch sử import (từ logs hoặc database)
    """
    # TODO: Implement import history tracking
    # Có thể tạo bảng import_logs để track
    
    return jsonify({
        'message': 'Import history feature coming soon',
        'imports': []
    }), 200


@import_bp.route('/bulk-delete', methods=['DELETE'])
@token_required
def bulk_delete_schedules(current_user):
    """
    Xóa nhiều lịch cùng lúc (sau khi import nhầm)
    
    Request Body:
    {
        "schedule_ids": [1, 2, 3, 4]
    }
    """
    user_id = current_user['user_id']
    data = request.get_json()
    
    if not data or 'schedule_ids' not in data:
        return jsonify({'error': 'Missing schedule_ids'}), 400
    
    schedule_ids = data['schedule_ids']
    
    if not isinstance(schedule_ids, list) or len(schedule_ids) == 0:
        return jsonify({'error': 'schedule_ids must be a non-empty list'}), 400
    
    # Security: Limit bulk delete
    if len(schedule_ids) > 100:
        return jsonify({'error': 'Cannot delete more than 100 schedules at once'}), 400
    
    with get_db() as conn:
        cursor = conn.cursor()
        
        # Delete only schedules belonging to user
        placeholders = ','.join('?' * len(schedule_ids))
        cursor.execute(f'''
            DELETE FROM schedule
            WHERE id IN ({placeholders})
            AND user_id = ?
        ''', schedule_ids + [user_id])
        
        deleted_count = cursor.rowcount
    
    return jsonify({
        'message': f'Deleted {deleted_count} schedules',
        'deleted_count': deleted_count
    }), 200


@import_bp.route('/supported-formats', methods=['GET'])
def get_supported_formats():
    """
    Trả về thông tin các format file được support
    """
    return jsonify({
        'supported_formats': [
            {
                'extension': 'xlsx',
                'name': 'Microsoft Excel (2007+)',
                'description': 'Recommended format',
                'max_size_mb': 5
            },
            {
                'extension': 'xls',
                'name': 'Microsoft Excel (Legacy)',
                'description': 'Old Excel format',
                'max_size_mb': 5
            },
            {
                'extension': 'csv',
                'name': 'Comma-Separated Values',
                'description': 'Simple text format',
                'max_size_mb': 5
            }
        ],
        'required_columns': [
            'Môn học (subject)',
            'Ngày (date)',
            'Giờ bắt đầu (start_time)',
            'Giờ kết thúc (end_time)'
        ],
        'optional_columns': [
            'Phòng (location)',
            'Loại (type)',
            'Mô tả (description)'
        ],
        'date_formats': [
            'YYYY-MM-DD (2025-10-20)',
            'DD/MM/YYYY (20/10/2025)',
            'DD-MM-YYYY (20-10-2025)'
        ],
        'time_formats': [
            'HH:MM (08:00)',
            'HH:MM:SS (08:00:00)',
            'HHhMM (8h00)'
        ]
    }), 200